from classes.readers.Reader import Reader
import openpyxl

class ExcellReader(Reader):

    def __init__(self, pathname):
        self.pathname = pathname
        self.elements = []
    
    def readFile(self, sheet_name):
        self.elements = []
        try:
            workbook = openpyxl.load_workbook(self.pathname, data_only=True)
            grades_sheet = workbook[sheet_name]
            num_of_cols = grades_sheet.max_column
            num_of_rows = grades_sheet.max_row
            attributes = []
            for i in range(1,num_of_cols+1):
                current_attribute = grades_sheet.cell(1,i)
                attributes.append(current_attribute.value)
            for i in range(2, num_of_rows + 1):
                still_valid = True
                current_student = {}
                for (index, attribute) in enumerate(attributes):
                    current_cell = grades_sheet.cell(i,index+1)
                    if(current_cell.value == None):
                        still_valid = False
                        break
                    current_student[attribute] = current_cell.value
                if not still_valid:
                    break
                self.elements.append(current_student)
        except:
            raise TypeError("Sheet name doesnt exists in file")

    def filterAttributes(self,attributes):
        filteredElements = []
        for element in self.elements:
            filteredElement = {} 
            for (key,value) in element.items():
                if key in attributes:
                    filteredElement[key] = value
            filteredElements.append(filteredElement)
        return filteredElements